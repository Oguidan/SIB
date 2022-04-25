from kivy.app import App
from openpyxl import *
from openpyxl.utils import *
wb = Workbook()
dest_filename = "fichier_SIB.xlsx"
ws1 = wb.active
ws1.title = "Feuille1"
for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Feuille2")
ws2["F5"] = 3.14
ws3 = wb.create_sheet(title="Feuille3")
for row in range (10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3["AA10"].value)
wb.save(filename=dest_filename)
# Remplir un groupe de cellule, par un exemple un ensemble de cellule de A2:C6
def Fill_groupe_of_cellula():
    input("Donner le nom de la feuille")